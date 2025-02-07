

from scipy import interpolate
import json
import numpy as np
import wx
# import requests
import base64
import io
import os
import pandas as pd
from libraries.ConfigFile import add_core_level_Data
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl import load_workbook
from libraries.Sheet_Operations import on_sheet_selected
from copy import deepcopy
# from Functions import convert_to_serializable_and_round


def save_all_sheets_with_plots(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    try:
        for sheet_name in window.Data['Core levels'].keys():
            # Select the sheet
            window.sheet_combobox.SetValue(sheet_name)
            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, sheet_name)

            # Get fitting data
            fit_data = window.get_data_for_save()

            # Save fitting data to Excel
            save_to_excel(window, fit_data, file_path, sheet_name)

            # Save plot to Excel
            save_plot_to_excel(window)

        # Save results table
        save_results_table(window)

        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        window.show_popup_message2("Save Complete", "All sheets, plots, and results table have been saved.")

    except Exception as e:
        wx.MessageBox(f"Error saving sheets with plots: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)
def save_data(window, data):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file path found in window.Data. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Save to Excel
        save_to_excel(window, data, file_path, sheet_name)

        save_plot_to_excel(window)

        # Save JSON file with entire window.Data
        json_file_path = os.path.splitext(file_path)[0] + '.json'

        # Create a copy of window.Data to modify
        json_data = window.Data.copy()

        # Convert numpy arrays and other non-serializable types to lists, and round floats
        json_data = convert_to_serializable_and_round(json_data)

        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # print(json.dumps(window.Data['Results']['Peak'], indent=2))
        print("Data Saved")
    except Exception as e:
        wx.MessageBox(f"Error saving data: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def convert_to_serializable_and_round2(obj, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            return {
                "rows": obj.GetNumberRows(),
                "cols": obj.GetNumberCols(),
                "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), decimal_places)
                          for col in range(obj.GetNumberCols())]
                         for row in range(obj.GetNumberRows())]
            }
        elif hasattr(obj, 'tolist'):  # This catches any object with a tolist method, like some lmfit results
            return convert_to_serializable_and_round(obj.tolist(), decimal_places)
        else:
            return obj
    except Exception as e:
        return str(obj)  # Return a string representation as a fallback

def convert_to_serializable_and_round(obj, decimal_places=2):
    try:
        if isinstance(obj, (float, np.float32, np.float64)):
            return round(float(obj), decimal_places)
        elif isinstance(obj, (int, np.int32, np.int64)):
            return int(obj)
        elif isinstance(obj, np.ndarray):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj.tolist()]
        elif isinstance(obj, list):
            return [convert_to_serializable_and_round(item, decimal_places) for item in obj]
        elif isinstance(obj, dict):
            return {k: convert_to_serializable_and_round(v, decimal_places) for k, v in obj.items()}
        elif isinstance(obj, wx.grid.Grid):
            if obj == window.results_grid:
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [{
                        "Label": f"Peak_{row}",
                        "Name": obj.GetCellValue(row, 0),
                        "Position": convert_to_serializable_and_round(obj.GetCellValue(row, 1), decimal_places),
                        "Height": convert_to_serializable_and_round(obj.GetCellValue(row, 2), decimal_places),
                        "FWHM": convert_to_serializable_and_round(obj.GetCellValue(row, 3), decimal_places),
                        "L/G": convert_to_serializable_and_round(obj.GetCellValue(row, 4), decimal_places),
                        "Area": convert_to_serializable_and_round(obj.GetCellValue(row, 5), decimal_places),
                        "at. %": convert_to_serializable_and_round(obj.GetCellValue(row, 6), decimal_places),
                        "Checkbox": obj.GetCellValue(row, 7),
                        "RSF": convert_to_serializable_and_round(obj.GetCellValue(row, 8), decimal_places),
                        "Fitting Model": obj.GetCellValue(row, 9),
                        "Rel. Area": convert_to_serializable_and_round(obj.GetCellValue(row, 10), decimal_places),
                        "Sigma": convert_to_serializable_and_round(obj.GetCellValue(row, 11), decimal_places),
                        "Gamma": convert_to_serializable_and_round(obj.GetCellValue(row, 12), decimal_places),
                        "Bkg Type": obj.GetCellValue(row, 13),
                        "Bkg Low": convert_to_serializable_and_round(obj.GetCellValue(row, 14), decimal_places),
                        "Bkg High": convert_to_serializable_and_round(obj.GetCellValue(row, 15), decimal_places),
                        "Bkg Offset Low": obj.GetCellValue(row, 16),
                        "Bkg Offset High": obj.GetCellValue(row, 17),
                        "Sheetname": obj.GetCellValue(row, 18),
                        "Pos. Constraint": obj.GetCellValue(row, 19),
                        "Height Constraint": obj.GetCellValue(row, 20),
                        "FWHM Constraint": obj.GetCellValue(row, 21),
                        "L/G Constraint": obj.GetCellValue(row, 22),
                        "Area Constraint": obj.GetCellValue(row, 23),
                        "Sigma Constraint": obj.GetCellValue(row, 24),
                        "Gamma Constraint": obj.GetCellValue(row, 25),
                        "Skew Constraint": obj.GetCellValue(row, 26)
                    } for row in range(obj.GetNumberRows())]
                }
            else:
                return {
                    "rows": obj.GetNumberRows(),
                    "cols": obj.GetNumberCols(),
                    "data": [[convert_to_serializable_and_round(obj.GetCellValue(row, col), decimal_places)
                              for col in range(obj.GetNumberCols())]
                             for row in range(obj.GetNumberRows())]
                }
        elif hasattr(obj, 'tolist'):
            return convert_to_serializable_and_round(obj.tolist(), decimal_places)
        else:
            return obj
    except Exception as e:
        return str(obj)

def convert_to_serializable(obj):
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    elif isinstance(obj, dict):
        return {k: convert_to_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_to_serializable(v) for v in obj]
    elif isinstance(obj, (np.int64, np.int32, np.float64, np.float32)):
        return obj.item()
    elif isinstance(obj, wx.grid.Grid):
        return {
            "rows": obj.GetNumberRows(),
            "cols": obj.GetNumberCols(),
            "data": [[obj.GetCellValue(row, col) for col in range(obj.GetNumberCols())]
                     for row in range(obj.GetNumberRows())]
        }
    else:
        return obj


def save_to_excel(window, data, file_path, sheet_name):
    existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Remove previously fitted data if it exists
    if existing_df.shape[1] > 5:
        existing_df = existing_df.iloc[:, :5]

    # Add columns if there aren't enough
    while existing_df.shape[1] < 5:
        existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

    # Ensure there's an empty column E
    if existing_df.shape[1] < 5:
        existing_df.insert(4, '', np.nan)

    if 'x_values' in data and data['x_values'] is not None:
        x_values = data['x_values'].to_numpy() if isinstance(data['x_values'], pd.Series) else data['x_values']
        filtered_data = pd.DataFrame({
            'BE': x_values
        })

        if data['background'] is not None and data['calculated_fit'] is not None:
            mask = np.isin(data['x_values'], x_values)
            y_values = data['y_values'][mask]

            if len(y_values) == len(data['calculated_fit']):
                residuals = y_values - data['calculated_fit']
                filtered_data['Residuals'] = residuals
            else:
                filtered_data['Residuals'] = np.nan
        else:
            filtered_data['Residuals'] = np.nan

        filtered_data['Background'] = data['background'] if data['background'] is not None else np.nan
        filtered_data['Calculated Fit'] = data['calculated_fit'] if data['calculated_fit'] is not None else np.nan

        if data['individual_peak_fits']:
            num_rows = len(x_values)
            num_peaks = data['peak_params_grid'].GetNumberRows() // 2
            for i in range(num_peaks):
                row = i * 2
                peak_label = data['peak_params_grid'].GetCellValue(row, 1)
                if i < len(data['individual_peak_fits']):
                    reversed_peak = np.array(data['individual_peak_fits'][i]) [::-1]
                    trimmed_peak = np.roll(reversed_peak, -1)[:num_rows]
                    filtered_data[peak_label] = trimmed_peak

        # Rename columns to avoid conflicts before inserting them
        for i, col in enumerate(filtered_data.columns):
            new_col_name = col
            while new_col_name in existing_df.columns:
                new_col_name += '_new'
            if col != 'Derivative':
                existing_df.insert(5 + i, new_col_name, filtered_data[col])

    # Ensure there are at least 23 columns (A to W)
    while existing_df.shape[1] < 23:
        existing_df[f'Column_{existing_df.shape[1] + 1}'] = ''

    # Rename columns starting with "Unnamed" or "Column" to empty string
    existing_df.columns = ['' if col.startswith(('Unnamed', 'Column')) else col for col in existing_df.columns]

    # Ensure column E is empty
    if existing_df.columns[4] != '':
        existing_df.rename(columns={existing_df.columns[4]: ''}, inplace=True)

    # Create DataFrame for peak fitting parameters
    peak_params_df = pd.DataFrame()
    for col in range(window.peak_params_grid.GetNumberCols()):
        col_name = window.peak_params_grid.GetColLabelValue(col)
        col_data = [window.peak_params_grid.GetCellValue(row, col) for row in
                    range(window.peak_params_grid.GetNumberRows())]
        peak_params_df[col_name] = col_data

    # Add peak_params_df to existing_df starting from column 23 (X)
    for i, col in enumerate(peak_params_df.columns):
        existing_df.insert(23 + i, col, peak_params_df[col])

    # Handle D-parameter derivative data
    if window.selected_fitting_method == "D-parameter":
        if 'Fitting' in window.Data['Core levels'][sheet_name] and 'Peaks' in window.Data['Core levels'][sheet_name][
            'Fitting']:
            d_param_data = window.Data['Core levels'][sheet_name]['Fitting']['Peaks'].get(
                'D-parameter')
            if d_param_data and 'Derivative' in d_param_data:
                filtered_data['Derivative'] = d_param_data['Derivative']
                existing_df.insert(7, 'Derivative', d_param_data['Derivative'])

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        existing_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Remove border from first row
        workbook = writer.book
        worksheet = workbook[sheet_name]
        for cell in worksheet[1]:
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style=None),
                right=openpyxl.styles.Side(style=None),
                top=openpyxl.styles.Side(style=None),
                bottom=openpyxl.styles.Side(style=None)
            )

            # Define styles
            thin_side = Side(style='thin')
            thick_side = Side(style='medium')
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
            bold_font = Font(bold=True)

            start_row = 2  # Assuming data starts from the second row
            num_peak_rows = window.peak_params_grid.GetNumberRows()
            end_row = start_row + num_peak_rows - 1
            start_col = 24  # Column X (24th column)
            end_col = worksheet.max_column

            for row in range(start_row - 1, end_row + 1):  # Start from header row
                for col in range(start_col, end_col + 1):
                    cell = worksheet.cell(row=row, column=col)

                    # Default to thin borders
                    left = right = top = bottom = thin_side

                    # Header row
                    if row == start_row - 1:
                        cell.fill = green_fill
                        cell.font = bold_font
                        top = thick_side

                    # Add thick borders for outer edges
                    if row == start_row - 1 or row == end_row:
                        bottom = thick_side
                    if col == start_col:
                        left = thick_side
                    if col == end_col:
                        right = thick_side

                    # Add thick bottom border for every second row (constraints row)
                    if (row - start_row + 1) % 2 == 0:
                        bottom = thick_side

                    cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # After saving to Excel, update the plot with the current limits
    if hasattr(window, 'plot_config'):
        limits = window.plot_config.get_plot_limits(window, sheet_name)
        window.ax.set_xlim(limits['Xmax'], limits['Xmin'])  # Reverse X-axis
        window.ax.set_ylim(limits['Ymin'], limits['Ymax'])
        window.canvas.draw_idle()


def refresh_sheets(window, on_sheet_selected_func):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file currently open. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    current_sheet = window.sheet_combobox.GetValue()
    file_path = window.Data['FilePath']

    try:
        # Save current state to JSON
        json_file_path = os.path.splitext(file_path)[0] + '.json'
        json_data = convert_to_serializable_and_round(window.Data)
        with open(json_file_path, 'w') as json_file:
            json.dump(json_data, json_file, indent=2)

        # Reopen the XLSX file
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        # Update sheet names in the combobox
        window.sheet_combobox.Clear()
        window.sheet_combobox.AppendItems(sheet_names)

        # Update window.Data with new sheet information
        for sheet_name in sheet_names:
            if sheet_name not in window.Data['Core levels']:
                window.Data = add_core_level_Data(window.Data, window, file_path, sheet_name)

        # Remove any sheets from window.Data that no longer exist in the Excel file
        sheets_to_remove = set(window.Data['Core levels'].keys()) - set(sheet_names)
        for sheet_name in sheets_to_remove:
            del window.Data['Core levels'][sheet_name]

        # Update the number of core levels
        window.Data['Number of Core levels'] = len(sheet_names)

        # Update B.E. and Raw Data with current Excel data
        for sheet_name in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            window.Data['Core levels'][sheet_name]['B.E.'] = df.iloc[:, 0].tolist()
            window.Data['Core levels'][sheet_name]['Raw Data'] = df.iloc[:, 1].tolist()

        # Set the current sheet as selected if it still exists, otherwise select the first sheet
        if current_sheet in sheet_names:
            window.sheet_combobox.SetValue(current_sheet)
        elif sheet_names:
            window.sheet_combobox.SetValue(sheet_names[0])
            current_sheet = sheet_names[0]

        # Update the plot for the current sheet
        event = wx.CommandEvent(wx.EVT_COMBOBOX.typeId)
        event.SetString(current_sheet)
        on_sheet_selected_func(window, event)

        # Update plot limits
        window.plot_config.update_plot_limits(window, current_sheet)

        # Refresh the plot
        window.plot_manager.plot_data(window)
        window.clear_and_replot()

        wx.MessageBox(f"Sheets refreshed. Total sheets: {len(sheet_names)}", "Success", wx.OK | wx.ICON_INFORMATION)

    except Exception as e:
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error refreshing sheets: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_to_json(window, file_path):
    json_file_path = os.path.splitext(file_path)[0] + '.json'

    # Create a copy of window.Data to modify
    data_to_save = window.Data.copy()

    def convert_to_list(item):
        if isinstance(item, np.ndarray):
            return item.tolist()
        elif isinstance(item, list):
            return item
        else:
            return item  # Return as is if it's neither numpy array nor list

    # Convert numpy arrays to lists for JSON serialization
    for sheet_name, sheet_data in data_to_save['Core levels'].items():
        sheet_data['B.E.'] = convert_to_list(sheet_data['B.E.'])
        sheet_data['Raw Data'] = convert_to_list(sheet_data['Raw Data'])
        sheet_data['Background']['Bkg Y'] = convert_to_list(sheet_data['Background']['Bkg Y'])

        if 'Fitting' in sheet_data and 'Peaks' in sheet_data['Fitting']:
            for peak_label, peak_data in sheet_data['Fitting']['Peaks'].items():
                if 'Y' in peak_data:
                    peak_data['Y'] = convert_to_list(peak_data['Y'])

    with open(json_file_path, 'w') as json_file:
        json.dump(data_to_save, json_file, indent=2)


def save_plot_to_excel(window):
   if 'FilePath' not in window.Data or not window.Data['FilePath']:
       wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
       return

   file_path = window.Data['FilePath']
   sheet_name = window.sheet_combobox.GetValue()
   is_survey = "survey" in sheet_name.lower() or "wide" in sheet_name.lower()

   try:
       # Get dimensions based on plot type
       width = window.survey_excel_width if is_survey else window.excel_width
       height = window.survey_excel_height if is_survey else window.excel_height
       dpi = window.survey_excel_dpi if is_survey else window.excel_dpi

       print("Save plot to Excel")

       # Save figure to buffer
       buf = io.BytesIO()
       original_size = window.figure.get_size_inches()
       window.figure.set_size_inches(width, height)
       window.figure.savefig(buf, format='png', dpi=dpi, bbox_inches='tight')
       window.figure.set_size_inches(original_size)
       buf.seek(0)

       # Save to Excel
       wb = openpyxl.load_workbook(file_path)
       ws = wb.create_sheet(sheet_name) if sheet_name not in wb.sheetnames else wb[sheet_name]

       # Clear existing images
       for img in ws._images:
           ws._images.remove(img)

       # Add new image
       img = Image(buf)
       ws.add_image(img, 'D6')
       wb.save(file_path)

       print(f"Plot saved to Excel file: {file_path}, Sheet: {sheet_name}")
       window.show_popup_message2("Plot saved into Excel file", f"Under sheet: {sheet_name}")

   except Exception as e:
       import traceback
       traceback.print_exc()
       wx.MessageBox(f"Error saving plot to Excel: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_png(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate PNG filename based on the Excel filename and sheet name
        png_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.png"
        png_filepath = os.path.join(os.path.dirname(file_path), png_filename)

        # Save the figure as PNG
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(png_filepath, format='png', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as PNG: {png_filepath}")
        window.show_popup_message2("Plot saved as PNG", f"File: {png_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as PNG: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_pdf(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate PDF filename based on the Excel filename and sheet name
        pdf_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.pdf"
        pdf_filepath = os.path.join(os.path.dirname(file_path), pdf_filename)

        # Save the figure as PDF
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(pdf_filepath, format='pdf', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as PDF: {pdf_filepath}")
        window.show_popup_message2("Plot saved as PDF", f"File: {pdf_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as PDF: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_plot_as_svg(window):
    # Check if a file is currently open
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()

    try:
        # Generate SVG filename based on the Excel filename and sheet name
        svg_filename = f"{os.path.splitext(os.path.basename(file_path))[0]}_{sheet_name}.svg"
        svg_filepath = os.path.join(os.path.dirname(file_path), svg_filename)

        # Save the figure as SVG
        original_size = window.figure.get_size_inches()
        window.figure.set_size_inches(window.export_width, window.export_height)
        window.figure.savefig(svg_filepath, format='svg', dpi=window.export_dpi, bbox_inches='tight')
        window.figure.set_size_inches(original_size)

        print(f"Plot saved as SVG: {svg_filepath}")
        window.show_popup_message2("Plot saved as SVG", f"File: {svg_filename}")

    except Exception as e:
        # Handle any errors that occur during saving
        import traceback
        traceback.print_exc()
        wx.MessageBox(f"Error saving plot as SVG: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)



def save_results_table(window):
    if 'FilePath' not in window.Data or not window.Data['FilePath']:
        wx.MessageBox("No file selected. Please open a file first.", "Error", wx.OK | wx.ICON_ERROR)
        return

    file_path = window.Data['FilePath']

    try:
        wb = openpyxl.load_workbook(file_path)

        sheet_name = 'Results Table'
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        def get_column_letter(n):
            result = ""
            while n > 0:
                n, remainder = divmod(n - 1, 26)
                result = chr(65 + remainder) + result
            return result

        headers = [window.results_grid.GetColLabelValue(col) for col in range(window.results_grid.GetNumberCols())]
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(window.results_grid.GetNumberRows()):
            for col in range(window.results_grid.GetNumberCols()):
                cell = ws.cell(row=row + 3, column=col + 2, value=window.results_grid.GetCellValue(row, col))
                cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                              bottom=Side(style='medium'))

        max_row = ws.max_row
        max_col = len(headers) + 1
        end_letter = get_column_letter(max_col)

        for row in ws[f'B2:{end_letter}2']:
            for cell in row:
                cell.border = thick_border

        for row in ws[f'B3:{end_letter}{max_row}']:
            for cell in row:
                cell.border = thin_border

        for col in range(2, max_col + 1):
            ws.cell(row=2, column=col).border = Border(left=ws.cell(row=2, column=col).border.left,
                                                       right=ws.cell(row=2, column=col).border.right,
                                                       top=Side(style='medium'),
                                                       bottom=ws.cell(row=2, column=col).border.bottom)
            ws.cell(row=max_row, column=col).border = Border(left=ws.cell(row=max_row, column=col).border.left,
                                                             right=ws.cell(row=max_row, column=col).border.right,
                                                             top=ws.cell(row=max_row, column=col).border.top,
                                                             bottom=Side(style='medium'))

        for row in range(2, max_row + 1):
            ws.cell(row=row, column=2).border = Border(left=Side(style='medium'),
                                                       right=ws.cell(row=row, column=2).border.right,
                                                       top=ws.cell(row=row, column=2).border.top,
                                                       bottom=ws.cell(row=row, column=2).border.bottom)
            ws.cell(row=row, column=max_col).border = Border(left=ws.cell(row=row, column=max_col).border.left,
                                                             right=Side(style='medium'),
                                                             top=ws.cell(row=row, column=max_col).border.top,
                                                             bottom=ws.cell(row=row, column=max_col).border.bottom)

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        wb.save(file_path)

        window.show_popup_message2("Table Saved","Results table has been saved to the Excel file.")

    except Exception as e:
        wx.MessageBox(f"Error saving results table: {str(e)}", "Error", wx.OK | wx.ICON_ERROR)


def save_state(window):
    current_sheet = window.sheet_combobox.GetValue()
    sheets_data = {}

    for sheet in window.Data['Core levels'].keys():
        if sheet == current_sheet:
            sheets_data[sheet] = {
                'peak_params_grid': get_grid_data(window.peak_params_grid),
                'peak_count': window.peak_count,
                'selected_peak_index': window.selected_peak_index
            }

    state = {
        'Data': deepcopy(window.Data),
        'current_sheet': current_sheet,
        'sheets': sheets_data,
        'results_grid': {
            'data': get_grid_data(window.results_grid),
            'num_rows': window.results_grid.GetNumberRows(),
            'checkbox_states': [window.results_grid.GetCellValue(row, 7)
                              for row in range(window.results_grid.GetNumberRows())]
        },
        'be_correction': window.be_correction  # Add BE correction to state
    }

    window.history.append(state)
    if len(window.history) > window.max_history:
        window.history.pop(0)
    window.redo_stack.clear()
    update_undo_redo_state(window)

def undo(window):
    if len(window.history) > 1:
        current_state = window.history.pop()
        window.redo_stack.append(current_state)
        previous_state = window.history[-1]
        restore_state(window, previous_state)
        update_undo_redo_state(window)

def redo(window):
    if window.redo_stack:
        next_state = window.redo_stack.pop()
        window.history.append(next_state)
        restore_state(window, next_state)
        update_undo_redo_state(window)


def restore_state(window, state):
    window.Data = deepcopy(state['Data'])
    window.be_correction = state.get('be_correction', 0)  # Restore BE correction with default 0
    window.be_correction_spinbox.SetValue(window.be_correction)

    # Restore sheet-specific data
    for sheet, sheet_data in state['sheets'].items():
        if sheet == state['current_sheet']:
            set_grid_data(window.peak_params_grid, sheet_data['peak_params_grid'])
            window.peak_count = sheet_data['peak_count']
            window.selected_peak_index = sheet_data['selected_peak_index']

    # Restore results grid
    results_grid_data = state['results_grid']
    current_rows = window.results_grid.GetNumberRows()
    target_rows = results_grid_data['num_rows']

    if current_rows < target_rows:
        window.results_grid.AppendRows(target_rows - current_rows)
    elif current_rows > target_rows:
        window.results_grid.DeleteRows(target_rows, current_rows - target_rows)

    set_grid_data(window.results_grid, results_grid_data['data'])

    # Restore checkbox states
    for row, checkbox_state in enumerate(results_grid_data['checkbox_states']):
        window.results_grid.SetCellRenderer(row, 7, wx.grid.GridCellBoolRenderer())
        window.results_grid.SetCellEditor(row, 7, wx.grid.GridCellBoolEditor())
        window.results_grid.SetCellValue(row, 7, checkbox_state)

    # Switch to the correct sheet
    window.sheet_combobox.SetValue(state['current_sheet'])
    on_sheet_selected(window, state['current_sheet'])

    window.clear_and_replot()

def get_grid_data(grid):
    data = []
    for row in range(grid.GetNumberRows()):
        row_data = [grid.GetCellValue(row, col) for col in range(grid.GetNumberCols())]
        data.append(row_data)
    return data

def set_grid_data(grid, data):
    grid.ClearGrid()
    if len(data) > grid.GetNumberRows():
        grid.AppendRows(len(data) - grid.GetNumberRows())
    for row, row_data in enumerate(data):
        for col, value in enumerate(row_data):
            grid.SetCellValue(row, col, value)


def update_undo_redo_state(window):
    can_undo = len(window.history) > 1
    can_redo = len(window.redo_stack) > 0

    # Update toolbar buttons if they exist
    if hasattr(window, 'undo_tool'):
        window.toolbar.EnableTool(window.undo_tool.GetId(), can_undo)
    if hasattr(window, 'redo_tool'):
        window.toolbar.EnableTool(window.redo_tool.GetId(), can_redo)

    # Update menu items if they exist
    if hasattr(window, 'edit_menu'):
        window.edit_menu.Enable(wx.ID_UNDO, can_undo)
        window.edit_menu.Enable(wx.ID_REDO, can_redo)




def save_plot_data_and_script(window, png_filepath):
    data = window.get_data_for_save()

    x_values = np.array(data['x_values'])
    y_values = np.array(data['y_values'])
    background = np.array(data['background'])
    envelope = np.array(data['calculated_fit'])
    fitted_peaks = [np.array(peak) for peak in data['individual_peak_fits']]

    # Ensure all arrays have the same length as x_values
    length = len(x_values)
    y_values = y_values[:length]
    background = background[:length]
    envelope = envelope[:length]
    fitted_peaks = [peak[:length] for peak in fitted_peaks]

    # Calculate full residuals
    full_residuals = y_values - envelope

    py_filepath = os.path.splitext(png_filepath)[0] + '.py'

    with open(py_filepath, 'w') as f:
        f.write("import numpy as np\n")
        f.write("import matplotlib.pyplot as plt\n\n")

        f.write(f"x_values = np.{repr(x_values)}\n")
        f.write(f"y_values = np.{repr(y_values)}\n")
        f.write(f"background = np.{repr(background)}\n")
        f.write(f"full_residuals = np.{repr(full_residuals)}\n")
        f.write(f"envelope = np.{repr(envelope)}\n")

        for i, peak in enumerate(fitted_peaks):
            f.write(f"fitted_peak_{i} = np.{repr(peak)}\n")

        f.write("\nplt.figure(figsize=(10, 6))\n")
        f.write("plt.scatter(x_values, y_values, color='black', s=10, label='Raw Data')\n")
        f.write("plt.plot(x_values, background, color='gray', linestyle='--', label='Background')\n")
        f.write("plt.plot(x_values, envelope, color='blue', label='Envelope')\n")
        f.write("plt.plot(x_values, full_residuals + np.max(y_values), color='green', label='Residuals')\n")

        for i in range(len(fitted_peaks)):
            f.write(
                f"plt.fill_between(x_values, background, fitted_peak_{i} + background, alpha=0.3, label='Fitted Peak {i + 1}')\n")
            f.write(f"plt.plot(x_values, fitted_peak_{i} + background, color='red', linestyle='-')\n")

        f.write("\nplt.xlabel('Binding Energy (eV)')\n")
        f.write("plt.ylabel('Intensity (CTS)')\n")
        f.write("plt.title('XPS Spectrum')\n")
        f.write("plt.legend()\n")
        f.write("plt.gca().invert_xaxis()\n")
        f.write("plt.tight_layout()\n")
        f.write("plt.show()\n")

    print(f"Plot data and script saved to {py_filepath}")


def create_plot_script_from_excel(window):
    file_path = window.Data['FilePath']
    sheet_name = window.sheet_combobox.GetValue()
    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Extract data
    x_values = df.iloc[:, 0].values  # Assuming BE is in the first column
    y_values = df.iloc[:, 1].values  # Assuming Raw Data is in the second column
    background = df['Background'].values if 'Background' in df.columns else None
    envelope = df['Calculated Fit'].values if 'Calculated Fit' in df.columns else None
    residuals = df['Residuals'].values if 'Residuals' in df.columns else None

    # Extract fitted peaks
    fitted_peaks = []
    peak_names = []
    calculated_fit_index = df.columns.get_loc('Calculated Fit') if 'Calculated Fit' in df.columns else -1
    for col in df.columns[calculated_fit_index + 1:]:
        if df[col].notna().any():  # Check if the column is not empty
            fitted_peaks.append(df[col].values)
            peak_names.append(col)
        else:
            break  # Stop when we find an empty column

    # Create Python script
    py_filepath = os.path.splitext(file_path)[0] + f'_{sheet_name}.py'

    with open(py_filepath, 'w') as f:
        f.write("import numpy as np\n")
        f.write("import matplotlib.pyplot as plt\n\n")

        f.write(f"x_values = np.{repr(x_values)}\n")
        f.write(f"y_values = np.{repr(y_values)}\n")
        if background is not None:
            f.write(f"background = np.{repr(background)}\n")
        if envelope is not None:
            f.write(f"envelope = np.{repr(envelope)}\n")
        if residuals is not None:
            f.write(f"residuals = np.{repr(residuals)}\n")

        for i, peak in enumerate(fitted_peaks):
            f.write(f"fitted_peak_{i+1} = np.{repr(peak)}\n")

        f.write("\nplt.figure(figsize=(8, 8))\n")

        if residuals is not None:
            f.write("plt.plot(x_values, residuals + np.max(y_values), color='green', label='Residuals')\n")

        for i, peak_name in enumerate(peak_names):
            f.write(f"plt.fill_between(x_values, background, fitted_peak_{i+1}, alpha=0.6, label='{peak_name}')\n")

        if envelope is not None:
            f.write("plt.plot(x_values, envelope, color='black', label='Envelope')\n")
        if background is not None:
            f.write("plt.plot(x_values, background, color='gray', linestyle='--', label='Background')\n")
        f.write("plt.scatter(x_values, y_values, color='black', s=20, label='Raw Data')\n")

        f.write("\nplt.xlabel('Binding Energy (eV)')\n")
        f.write("plt.ylabel('Intensity (CTS)')\n")
        f.write(f"plt.title('XPS Spectrum - {sheet_name}')\n")
        f.write("plt.legend(loc='upper left')\n")
        f.write(f"plt.xlim({max(x_values)}, {min(x_values)})\n")  # Reverse x-axis
        f.write("plt.tight_layout()\n")
        f.write("plt.show()\n")

    print(f"Plot script created: {py_filepath}")

def save_peaks_library(window):
    sheet_name = window.sheet_combobox.GetValue()
    with wx.FileDialog(window, "Save peaks library", wildcard="JSON files (*.json)|*.json",
                       defaultDir="Peaks Library", style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        path = fileDialog.GetPath()

        # Convert numpy arrays to lists before saving
        def convert_numpy_to_list(obj):
            if isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, dict):
                return {k: convert_numpy_to_list(v) for k, v in obj.items()}
            elif isinstance(obj, list):
                return [convert_numpy_to_list(item) for item in obj]
            return obj

        peaks_data = {
            'Core levels': {
                sheet_name: {
                    'Fitting': convert_numpy_to_list(window.Data['Core levels'][sheet_name]['Fitting'])
                }
            }
        }

        with open(path, 'w') as f:
            json.dump(peaks_data, f, indent=2)

def load_peaks_library(window):
    import wx
    import json
    from libraries.Sheet_Operations import on_sheet_selected
    from libraries.Fitting_Screen import FittingWindow

    with wx.FileDialog(window, "Load peaks library", wildcard="JSON files (*.json)|*.json",
                       defaultDir="Peaks Library", style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
        if fileDialog.ShowModal() == wx.ID_CANCEL:
            return
        path = fileDialog.GetPath()

    with open(path) as f:
        peaks_data = json.load(f)

    sheet_name = window.sheet_combobox.GetValue()
    source_sheet = list(peaks_data['Core levels'].keys())[0]
    source_data = peaks_data['Core levels'][source_sheet]

    # Copy only fitting data
    window.Data['Core levels'][sheet_name]['Fitting'] = source_data['Fitting']

    # Update display
    on_sheet_selected(window, sheet_name)

    # Ensure `fitting_window` exists and call `on_background`
    if not hasattr(window, 'fitting_window') or window.fitting_window is None:
        window.fitting_window = FittingWindow(parent=window)  # Initialize if necessary

    bg_low = window.peak_params_grid.GetCellValue(0, 15)
    bg_high = window.peak_params_grid.GetCellValue(0, 16)

    if bg_low and bg_high:
        window.bg_min_energy = float(bg_low)
        window.bg_max_energy = float(bg_high)
    else:
        x_values = window.Data['Core levels'][sheet_name]['B.E.']
        window.bg_min_energy = min(x_values) + 0.2
        window.bg_max_energy = max(x_values) - 0.2

    # window.fitting_window.on_background(None)



def save_peaks_to_github(window, filename):
    API_URL = "https://api.github.com/repos/KherveFitting/peaks-library/contents/peaks"

    sheet_name = window.sheet_combobox.GetValue()
    peaks_data = {
        'Core levels': {
            sheet_name: {
                'Fitting': window.Data['Core levels'][sheet_name]['Fitting'],
                'Background': window.Data['Core levels'][sheet_name]['Background']
            }
        }
    }

    content = base64.b64encode(json.dumps(peaks_data).encode()).decode()

    headers = {
        "Accept": "application/vnd.github.v3+json"
    }

    # List existing files to check if file exists
    try:
        response = requests.get(f"{API_URL}/{filename}.json", headers=headers)
        response.raise_for_status()
    except requests.exceptions.RequestException:
        wx.MessageBox("Error accessing peaks library. Check your internet connection.", "Error")
        return
def load_peaks_from_github(window):
    API_URL = "https://api.github.com/repos/KherveFitting/peaks-library/contents/peaks"

    try:
        # Get list of available peak libraries
        response = requests.get(API_URL)
        response.raise_for_status()
        files = [f['name'] for f in response.json() if f['name'].endswith('.json')]

        # Show file selection dialog
        dialog = wx.SingleChoiceDialog(window, "Select a peak library to load:",
                                       "Load Peak Library", files)
        if dialog.ShowModal() == wx.ID_OK:
            filename = dialog.GetStringSelection()

            # Get file content
            response = requests.get(f"{API_URL}/{filename}")
            response.raise_for_status()
            content = base64.b64decode(response.json()['content']).decode()
            peaks_data = json.loads(content)

            # Rest of your existing load code...
            sheet_name = window.sheet_combobox.GetValue()
            source_sheet = list(peaks_data['Core levels'].keys())[0]
            source_data = peaks_data['Core levels'][source_sheet]

            window.Data['Core levels'][sheet_name]['Fitting'] = source_data['Fitting']
            if 'Background' in source_data:
                window.Data['Core levels'][sheet_name]['Background'] = source_data['Background']
                window.background = np.array(source_data['Background']['Bkg Y'])
                window.bg_min_energy = float(source_data['Background'].get('Bkg Low', 0))
                window.bg_max_energy = float(source_data['Background'].get('Bkg High', 0))
                window.background_method = str(source_data['Background'].get('Bkg Type', 'Smart'))
                window.offset_h = float(source_data['Background'].get('Bkg Offset High', 0))
                window.offset_l = float(source_data['Background'].get('Bkg Offset Low', 0))

            from libraries.Sheet_Operations import on_sheet_selected
            on_sheet_selected(window, sheet_name)

    except requests.exceptions.RequestException:
        wx.MessageBox("Error accessing peaks library. Check your internet connection.", "Error")